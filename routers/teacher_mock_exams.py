import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any
import re

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from db import get_db
import models
from utils.security import require_role

router = APIRouter(
    prefix="/teacher/mock-exams",
    tags=["teacher_mock_exams"],
)

MOCK_TYPE_BY_NUMBER = {
    1: "purpose",
    2: "mood",
    3: "claim",
    4: "implication",
    5: "gist",
    6: "topic",
    7: "title",
    8: "content_match",
    9: "grammar",
    10: "vocabulary",
    11: "blank",
    12: "blank",
    13: "blank",
    14: "blank",
    15: "irrelevant",
    16: "order",
    17: "order",
    18: "insertion",
    19: "insertion",
    20: "summary",
}

TYPE_LABELS = {
    "purpose": "목적",
    "mood": "심경/분위기",
    "claim": "주장",
    "implication": "함의",
    "gist": "요지",
    "topic": "주제",
    "title": "제목",
    "content_match": "일치",
    "grammar": "어법",
    "vocabulary": "어휘",
    "blank": "빈칸",
    "irrelevant": "무관한 문장",
    "order": "순서",
    "insertion": "삽입",
    "summary": "요약",
}

TYPE_ORDER = []
for number in sorted(MOCK_TYPE_BY_NUMBER):
    type_code = MOCK_TYPE_BY_NUMBER[number]
    if type_code not in TYPE_ORDER:
        TYPE_ORDER.append(type_code)


class MockExamCreateRequest(BaseModel):
    grade: str = Field(..., examples=["고2"])
    year: int = Field(..., examples=[2024])
    month: int = Field(..., examples=[9])
    title: str = Field(..., examples=["2024년 9월 고2 모의고사"])
    has_listening: bool = False


class MockQuestionUpsert(BaseModel):
    number: int
    question_type: str
    source: str | None = None
    passage: str | None = None
    question_text: str
    options: list[str]
    answer_index: int
    explanation: str | None = None
    passage_group_id: str | None = None


class MockQuestionsUpsertRequest(BaseModel):
    questions: list[MockQuestionUpsert]


class MockQuestionPatchRequest(BaseModel):
    source: str | None = None
    passage: str | None = None
    question_text: str | None = None
    options: list[str] | None = None
    answer: int | None = None
    answer_index: int | None = None
    explanation: str | None = None
    passage_group_id: str | None = None


REQUIRED_UPLOAD_COLUMNS = [
    "number",
    "question_type",
    "question_text",
    "option_1",
    "option_2",
    "option_3",
    "option_4",
    "option_5",
    "answer",
]

OPTIONAL_UPLOAD_COLUMNS = [
    "source",
    "passage",
    "explanation",
    "passage_group",
]

COLUMN_ALIASES = {
    "number": "number",
    "no": "number",
    "번호": "number",
    "문항번호": "number",
    "문제번호": "number",
    "source": "source",
    "출처": "source",
    "자료출처": "source",
    "question_type": "question_type",
    "questiontype": "question_type",
    "유형": "question_type",
    "문제유형": "question_type",
    "문항유형": "question_type",
    "passage": "passage",
    "passagetext": "passage",
    "passage_text": "passage",
    "지문": "passage",
    "본문": "passage",
    "question_text": "question_text",
    "questiontext": "question_text",
    "문제": "question_text",
    "문항": "question_text",
    "발문": "question_text",
    "질문": "question_text",
    "option_1": "option_1",
    "option1": "option_1",
    "선택지1": "option_1",
    "보기1": "option_1",
    "1번": "option_1",
    "①": "option_1",
    "option_2": "option_2",
    "option2": "option_2",
    "선택지2": "option_2",
    "보기2": "option_2",
    "2번": "option_2",
    "②": "option_2",
    "option_3": "option_3",
    "option3": "option_3",
    "선택지3": "option_3",
    "보기3": "option_3",
    "3번": "option_3",
    "③": "option_3",
    "option_4": "option_4",
    "option4": "option_4",
    "선택지4": "option_4",
    "보기4": "option_4",
    "4번": "option_4",
    "④": "option_4",
    "option_5": "option_5",
    "option5": "option_5",
    "선택지5": "option_5",
    "보기5": "option_5",
    "5번": "option_5",
    "⑤": "option_5",
    "answer": "answer",
    "answer_index": "answer",
    "answerindex": "answer",
    "정답": "answer",
    "답": "answer",
    "explanation": "explanation",
    "해설": "explanation",
    "설명": "explanation",
    "passage_group": "passage_group",
    "passagegroup": "passage_group",
    "지문그룹": "passage_group",
    "그룹": "passage_group",
}

QUESTION_TYPE_ALIASES = {
    "purpose": "purpose",
    "목적": "purpose",
    "mood": "mood",
    "심경": "mood",
    "심경/분위기": "mood",
    "분위기": "mood",
    "claim": "claim",
    "주장": "claim",
    "implication": "implication",
    "함의": "implication",
    "gist": "gist",
    "요지": "gist",
    "topic": "topic",
    "주제": "topic",
    "title": "title",
    "제목": "title",
    "content_match": "content_match",
    "contentmatch": "content_match",
    "일치": "content_match",
    "내용일치": "content_match",
    "grammar": "grammar",
    "어법": "grammar",
    "vocabulary": "vocabulary",
    "어휘": "vocabulary",
    "blank": "blank",
    "빈칸": "blank",
    "irrelevant": "irrelevant",
    "무관한문장": "irrelevant",
    "무관문": "irrelevant",
    "order": "order",
    "순서": "order",
    "insertion": "insertion",
    "삽입": "insertion",
    "summary": "summary",
    "요약": "summary",
}

HEADER_SCAN_LIMIT = 30
HEADER_REQUIRED_THRESHOLD = 7
HEADER_CORE_COLUMNS = {
    "question_text",
    "option_1",
    "option_2",
    "option_3",
    "option_4",
    "option_5",
    "answer",
}
GUIDANCE_ROW_MARKERS = {
    "예시",
    "작성안내",
    "작성 안내",
    "필수 컬럼",
    "선택 컬럼",
    "업로드",
    "엑셀 작성 안내",
    "사용 방법",
}


def _user_id(current_user: dict[str, Any]) -> int:
    return int(current_user["sub"])


def _validate_exam_payload(payload: MockExamCreateRequest):
    if payload.grade not in {"고1", "고2", "고3"}:
        raise HTTPException(
            status_code=400,
            detail="grade must be one of 고1, 고2, 고3",
        )
    if payload.month < 1 or payload.month > 12:
        raise HTTPException(status_code=400, detail="month must be 1-12")


def _validate_question(question: MockQuestionUpsert):
    expected_type = MOCK_TYPE_BY_NUMBER.get(question.number)
    if expected_type is None:
        raise HTTPException(
            status_code=400,
            detail=f"number must be 1-20: {question.number}",
        )
    if question.question_type != expected_type:
        raise HTTPException(
            status_code=400,
            detail=(
                f"question {question.number} must use "
                f"question_type={expected_type}"
            ),
        )
    if len(question.options) != 5:
        raise HTTPException(
            status_code=400,
            detail=f"question {question.number} must have exactly 5 options",
        )
    if question.answer_index < 0 or question.answer_index > 4:
        raise HTTPException(
            status_code=400,
            detail=f"question {question.number} answer_index must be 0-4",
        )


def _get_teacher_question(
    db: Session,
    mock_exam_id: int,
    question_id: int,
    teacher_id: int,
):
    exam = _get_teacher_exam(db, mock_exam_id, teacher_id)
    question = (
        db.query(models.MockQuestion)
        .filter(
            models.MockQuestion.id == question_id,
            models.MockQuestion.mock_exam_id == exam.id,
        )
        .first()
    )
    if not question:
        raise HTTPException(status_code=404, detail="Mock question not found")
    return exam, question


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_column_key(value: Any) -> str:
    text = _clean_cell(value).replace("\ufeff", "")
    text = text.replace("\n", "").replace("\r", "").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = text.replace("-", "_")
    return text


def _canonical_column_name(value: Any) -> str:
    key = _normalize_column_key(value)
    return COLUMN_ALIASES.get(key, key)


def _build_header_map(raw_headers: list[Any]):
    received_columns = [_clean_cell(header) for header in raw_headers]
    header_map: dict[str, str] = {}
    canonical_headers: list[str] = []
    for raw_header in received_columns:
        if not raw_header:
            continue
        canonical = _canonical_column_name(raw_header)
        header_map[raw_header] = canonical
        if canonical not in canonical_headers:
            canonical_headers.append(canonical)
    return header_map, canonical_headers, received_columns


def _raise_missing_columns(missing: list[str], received_columns: list[str]):
    print("UPLOAD COLUMNS:", received_columns)
    print("MISSING COLUMNS:", missing)
    raise HTTPException(
        status_code=400,
        detail={
            "message": "Missing required columns",
            "missing_columns": missing,
            "received_columns": received_columns,
            "required_columns": REQUIRED_UPLOAD_COLUMNS,
            "optional_columns": OPTIONAL_UPLOAD_COLUMNS,
        },
    )


def _normalize_question_type(value: Any) -> str:
    text = _clean_cell(value)
    key = _normalize_column_key(text)
    return QUESTION_TYPE_ALIASES.get(key, text.strip().lower())


def _header_candidate(raw_headers: list[Any], row_number: int, sheet_name: str | None = None):
    received_columns = [_clean_cell(header) for header in raw_headers]
    mapped_columns = [
        _canonical_column_name(header) if header else ""
        for header in received_columns
    ]
    mapped_set = {column for column in mapped_columns if column}
    required_count = len(set(REQUIRED_UPLOAD_COLUMNS) & mapped_set)
    core_count = len(HEADER_CORE_COLUMNS & mapped_set)
    optional_count = len(set(OPTIONAL_UPLOAD_COLUMNS) & mapped_set)
    score = required_count * 10 + core_count * 3 + optional_count
    acceptable = (
        required_count >= HEADER_REQUIRED_THRESHOLD
        or HEADER_CORE_COLUMNS.issubset(mapped_set)
    )
    return {
        "sheet_name": sheet_name,
        "row_number": row_number,
        "received_columns": received_columns,
        "mapped_columns": mapped_columns,
        "required_count": required_count,
        "score": score,
        "acceptable": acceptable,
    }


def _best_header_candidate(
    raw_rows: list[list[Any]],
    sheet_name: str | None = None,
):
    best = None
    for index, row in enumerate(raw_rows[:HEADER_SCAN_LIMIT], start=1):
        candidate = _header_candidate(row, index, sheet_name)
        if best is None or candidate["score"] > best["score"]:
            best = candidate
    return best if best and best["acceptable"] else None


def _raise_header_detection_failed(first_rows: list[dict]):
    print("HEADER DETECTION FAILED")
    print("RECEIVED FIRST ROWS:", first_rows)
    raise HTTPException(
        status_code=400,
        detail={
            "message": (
                "엑셀에서 실제 컬럼명 행을 찾지 못했습니다. "
                "첫 번째 행 또는 안내문 아래에 필요한 컬럼을 넣어 주세요."
            ),
            "required_columns": REQUIRED_UPLOAD_COLUMNS,
            "column_examples": (
                "번호, 유형, 문제, 선택지1, 선택지2, 선택지3, 선택지4, 선택지5, 정답"
            ),
            "received_first_rows": first_rows,
        },
    )


def _is_guidance_row(data: dict[str, str]):
    values = " ".join(_clean_cell(value) for value in data.values()).strip()
    if not values:
        return True
    compact = values.replace(" ", "")
    return any(marker.replace(" ", "") in compact for marker in GUIDANCE_ROW_MARKERS)


def _has_minimum_data(data: dict[str, str]):
    return bool(
        _clean_cell(data.get("number"))
        and _clean_cell(data.get("question_text"))
        and _clean_cell(data.get("answer"))
    )


def _rich_text_to_underlined_html(value: Any) -> str:
    if value is None or isinstance(value, (str, int, float, bool)):
        return _clean_cell(value)

    try:
        parts = list(value)
    except TypeError:
        return _clean_cell(value)

    converted: list[str] = []
    has_rich_part = False
    for part in parts:
        text = getattr(part, "text", part)
        if text is None:
            continue
        font = getattr(part, "font", None)
        underline = getattr(font, "u", None) or getattr(font, "underline", None)
        if font is not None:
            has_rich_part = True
        converted.append(f"<u>{text}</u>" if underline else str(text))

    if not has_rich_part:
        return _clean_cell(value)
    return "".join(converted).strip()


def _cell_text(cell: Any) -> str:
    text = _rich_text_to_underlined_html(getattr(cell, "value", None))
    font = getattr(cell, "font", None)
    underline = getattr(font, "u", None) or getattr(font, "underline", None)
    if text and underline and "<u>" not in text.lower():
        return f"<u>{text}</u>"
    return text


def _read_csv_rows(raw: bytes):
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp949")
    raw_rows = list(csv.reader(StringIO(text)))
    if not raw_rows:
        raise HTTPException(
            status_code=400,
            detail=(
                "CSV header row is required. The first row must contain columns "
                "such as number, question_type, question_text, option_1, ..., answer."
            ),
        )

    candidate = _best_header_candidate(raw_rows, "CSV")
    if not candidate:
        first_rows = [
            {"row": index, "values": [_clean_cell(value) for value in row]}
            for index, row in enumerate(raw_rows[:5], start=1)
        ]
        _raise_header_detection_failed(first_rows)

    headers = candidate["mapped_columns"]
    received_columns = candidate["received_columns"]
    print("DETECTED HEADER ROW:", candidate["row_number"])
    print("UPLOAD COLUMNS BEFORE MAP:", received_columns)
    print("UPLOAD COLUMNS AFTER MAP:", headers)
    missing = [name for name in REQUIRED_UPLOAD_COLUMNS if name not in headers]
    if missing:
        _raise_missing_columns(missing, received_columns)

    rows = []
    data_start_index = int(candidate["row_number"])
    for index, raw_row in enumerate(raw_rows[data_start_index:], start=data_start_index + 1):
        if not any(_clean_cell(value) for value in raw_row):
            continue
        data = {}
        for column_index, value in enumerate(raw_row):
            if column_index < len(headers):
                canonical = headers[column_index]
                if not canonical:
                    continue
                data[canonical] = _clean_cell(value)
        if _is_guidance_row(data) or not _has_minimum_data(data):
            continue
        rows.append(
            {
                "row_number": index,
                "data": data,
            }
        )
    print("VALID ROW COUNT:", len(rows))
    return rows


def _read_xlsx_rows(raw: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required for Excel upload. Install requirements.txt.",
        ) from exc

    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, rich_text=True)
    except TypeError:
        workbook = load_workbook(BytesIO(raw), data_only=True)
    print("EXCEL SHEETS:", workbook.sheetnames)

    best_candidate = None
    first_rows: list[dict] = []
    for sheet in workbook.worksheets:
        print("CHECKING SHEET:", sheet.title)
        max_scan_row = min(sheet.max_row or 1, HEADER_SCAN_LIMIT)
        raw_rows = []
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_scan_row),
            start=1,
        ):
            values = [_clean_cell(cell.value) for cell in cells]
            raw_rows.append(values)
            if len(first_rows) < 8:
                first_rows.append({
                    "sheet": sheet.title,
                    "row": row_number,
                    "values": values,
                })

        candidate = _best_header_candidate(raw_rows, sheet.title)
        if candidate and (
            best_candidate is None or candidate["score"] > best_candidate["score"]
        ):
            best_candidate = candidate

    if not best_candidate:
        _raise_header_detection_failed(first_rows)

    sheet = workbook[best_candidate["sheet_name"]]
    headers = best_candidate["mapped_columns"]
    received_columns = best_candidate["received_columns"]
    header_row_number = int(best_candidate["row_number"])
    print("SELECTED SHEET:", best_candidate["sheet_name"])
    print("DETECTED HEADER ROW:", header_row_number)
    print("UPLOAD COLUMNS BEFORE MAP:", received_columns)
    print("UPLOAD COLUMNS AFTER MAP:", headers)
    print("UPLOAD COLUMNS:", received_columns)
    missing = [name for name in REQUIRED_UPLOAD_COLUMNS if name not in headers]
    if missing:
        _raise_missing_columns(missing, received_columns)

    rows = []
    for index, cells in enumerate(sheet.iter_rows(min_row=header_row_number + 1), start=header_row_number + 1):
        if not any(_cell_text(cell) for cell in cells):
            continue
        row = {
            headers[column_index]: _cell_text(cell)
            for column_index, cell in enumerate(cells)
            if column_index < len(headers) and headers[column_index]
        }
        if _is_guidance_row(row) or not _has_minimum_data(row):
            continue
        rows.append({"row_number": index, "data": row})
    print("VALID ROW COUNT:", len(rows))
    return rows


def _read_upload_rows(filename: str, raw: bytes):
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        return _read_csv_rows(raw)
    if suffix in {"xlsx", "xlsm"}:
        return _read_xlsx_rows(raw)
    raise HTTPException(
        status_code=400,
        detail="Only .csv, .xlsx, and .xlsm files are supported",
    )


def _parse_int(value: str, *, row_number: int, field_name: str, errors: list[str]):
    text = _clean_cell(value)
    try:
        return int(text)
    except (TypeError, ValueError):
        try:
            numeric_value = float(text)
            if numeric_value.is_integer():
                return int(numeric_value)
        except (TypeError, ValueError):
            pass
    errors.append(f"row {row_number}: {field_name} must be an integer")
    return 0


def _parse_answer(value: Any, *, row_number: int, errors: list[str]):
    text = _clean_cell(value)
    aliases = {
        "①": 1,
        "②": 2,
        "③": 3,
        "④": 4,
        "⑤": 5,
        "1번": 1,
        "2번": 2,
        "3번": 3,
        "4번": 4,
        "5번": 5,
    }
    if text in aliases:
        return aliases[text]

    match = re.search(r"[1-5]", text)
    if match:
        return int(match.group(0))

    errors.append(f"row {row_number}: answer must be 1-5")
    return 0


def _questions_from_upload_rows(rows: list[dict]):
    errors: list[str] = []
    questions: list[MockQuestionUpsert] = []
    seen_numbers: set[int] = set()
    parsed_numbers: set[int] = set()

    for row in rows:
        row_errors: list[str] = []
        row_number = row["row_number"]
        data = row["data"]

        number = _parse_int(
            data.get("number", ""),
            row_number=row_number,
            field_name="number",
            errors=row_errors,
        )
        answer = _parse_answer(
            data.get("answer", ""),
            row_number=row_number,
            errors=row_errors,
        )

        if 1 <= number <= 20:
            parsed_numbers.add(number)
            if number in seen_numbers:
                row_errors.append(
                    f"row {row_number}: duplicate question number {number}"
                )
            seen_numbers.add(number)

        options = [data.get(f"option_{index}", "") for index in range(1, 6)]
        missing_options = [
            f"option_{index}"
            for index, option in enumerate(options, start=1)
            if not option
        ]
        if missing_options:
            row_errors.append(
                f"row {row_number}: missing options {', '.join(missing_options)}"
            )

        question_type = _normalize_question_type(data.get("question_type", ""))
        expected_type = MOCK_TYPE_BY_NUMBER.get(number)
        if expected_type is None:
            row_errors.append(f"row {row_number}: number must be 1-20")
        elif question_type != expected_type:
            row_errors.append(
                f"row {row_number}: question_type must be {expected_type} "
                f"for number {number}"
            )

        if not data.get("question_text"):
            row_errors.append(f"row {row_number}: question_text is required")

        if row_errors:
            errors.extend(row_errors)
            continue

        questions.append(
            MockQuestionUpsert(
                number=number,
                question_type=question_type,
                source=data.get("source") or None,
                passage=data.get("passage") or None,
                question_text=data.get("question_text", ""),
                options=options,
                answer_index=answer - 1,
                explanation=data.get("explanation") or "해설 없음",
                passage_group_id=data.get("passage_group") or None,
            )
        )

    expected_numbers = set(range(1, 21))
    missing_numbers = sorted(expected_numbers - parsed_numbers)
    if missing_numbers:
        errors.append(
            "missing question numbers: "
            + ", ".join(str(number) for number in missing_numbers)
        )

    if errors:
        raise HTTPException(
            status_code=400,
            detail={"message": "Upload validation failed", "errors": errors},
        )

    return sorted(questions, key=lambda question: question.number)


def _serialize_question(question: models.MockQuestion, include_answer: bool = True):
    data = {
        "id": question.id,
        "mock_exam_id": question.mock_exam_id,
        "number": question.number,
        "question_type": question.question_type,
        "type_label": TYPE_LABELS.get(question.question_type, question.question_type),
        "source": question.source,
        "passage": question.passage,
        "question_text": question.question_text,
        "options": question.options or [],
        "passage_group_id": question.passage_group_id,
        "created_at": question.created_at.isoformat()
        if question.created_at
        else None,
    }
    if include_answer:
        data["answer_index"] = question.answer_index
        data["explanation"] = question.explanation
    return data


def _serialize_exam(exam: models.MockExam, include_questions: bool = False):
    questions = sorted(exam.questions or [], key=lambda item: item.number)
    data = {
        "id": exam.id,
        "grade": exam.grade,
        "year": exam.year,
        "month": exam.month,
        "title": exam.title,
        "total_questions": exam.total_questions,
        "total_score": exam.total_score,
        "has_listening": exam.has_listening,
        "created_by": exam.created_by,
        "created_at": exam.created_at.isoformat() if exam.created_at else None,
        "question_count": len(questions),
        "is_complete": len(questions) == exam.total_questions,
    }
    if include_questions:
        data["questions"] = [
            _serialize_question(question, include_answer=True)
            for question in questions
        ]
    return data


def _score_text(value: float | int | None):
    if value is None:
        return 0
    return round(float(value), 1)


def _type_label(question_type: str):
    return TYPE_LABELS.get(question_type, question_type)


def _question_type_counts(exam: models.MockExam):
    counts: dict[str, int] = {type_code: 0 for type_code in TYPE_ORDER}
    for question in exam.questions or []:
        counts[question.question_type] = counts.get(question.question_type, 0) + 1
    return counts


def _latest_attempts_by_student(exam: models.MockExam):
    latest: dict[int, models.MockAttempt] = {}
    attempts = sorted(
        exam.attempts or [],
        key=lambda item: item.submitted_at or item.started_at or datetime.min,
        reverse=True,
    )
    for attempt in attempts:
        if attempt.user_id not in latest:
            latest[attempt.user_id] = attempt
    return list(latest.values())


def _weak_labels_for_attempt(attempt: models.MockAttempt, questions):
    weak: list[str] = []
    seen: set[str] = set()
    answered_question_ids = set()
    for answer in sorted(attempt.answers or [], key=lambda item: item.id):
        answered_question_ids.add(answer.mock_question_id)
        if answer.is_correct:
            continue
        label = TYPE_LABELS.get(answer.question_type, answer.question_type)
        if label not in seen:
            weak.append(label)
            seen.add(label)

    for question in questions:
        if question.id in answered_question_ids:
            continue
        label = TYPE_LABELS.get(question.question_type, question.question_type)
        if label not in seen:
            weak.append(label)
            seen.add(label)
    return weak


def _teacher_mock_attempts(
    db: Session,
    teacher_id: int,
    student_id: int | None = None,
):
    query = (
        db.query(models.MockAttempt)
        .join(models.MockExam, models.MockAttempt.mock_exam_id == models.MockExam.id)
        .filter(models.MockExam.created_by == teacher_id)
    )
    if student_id is not None:
        query = query.filter(models.MockAttempt.user_id == student_id)
    return (
        query.order_by(
            models.MockAttempt.submitted_at.desc(),
            models.MockAttempt.id.desc(),
        )
        .all()
    )


def _teacher_mock_student_report(attempts: list[models.MockAttempt]):
    attempt_count = len(attempts)
    scores = [float(attempt.score or 0) for attempt in attempts]
    correct_by_type: dict[str, int] = {type_code: 0 for type_code in TYPE_ORDER}
    total_by_type: dict[str, int] = {type_code: 0 for type_code in TYPE_ORDER}

    for attempt in attempts:
        exam = attempt.mock_exam
        if exam:
            for type_code, count in _question_type_counts(exam).items():
                total_by_type[type_code] = total_by_type.get(type_code, 0) + count
        else:
            for answer in attempt.answers or []:
                total_by_type[answer.question_type] = (
                    total_by_type.get(answer.question_type, 0) + 1
                )

        for answer in attempt.answers or []:
            correct_by_type.setdefault(answer.question_type, 0)
            total_by_type.setdefault(answer.question_type, 0)
            if answer.is_correct:
                correct_by_type[answer.question_type] += 1

    type_stats = []
    for type_code in TYPE_ORDER:
        total = total_by_type.get(type_code, 0)
        correct = correct_by_type.get(type_code, 0)
        rate = round(correct / total * 100, 1) if total else 0
        type_stats.append(
            {
                "type": type_code,
                "label": _type_label(type_code),
                "correct": correct,
                "total": total,
                "rate": rate,
            }
        )

    weak_types = [
        item["label"]
        for item in sorted(
            [item for item in type_stats if item["total"] > 0],
            key=lambda item: (item["rate"], -item["total"]),
        )[:3]
        if item["rate"] < 70
    ]

    recent_attempts = []
    for attempt in attempts[:10]:
        exam = attempt.mock_exam
        recent_attempts.append(
            {
                "attempt_id": attempt.id,
                "mock_exam_id": attempt.mock_exam_id,
                "title": exam.title if exam else "삭제된 모의고사",
                "grade": exam.grade if exam else "-",
                "year": exam.year if exam else None,
                "month": exam.month if exam else None,
                "score": _score_text(attempt.score),
                "correct_count": attempt.correct_count,
                "total_questions": attempt.total_questions,
                "submitted_at": attempt.submitted_at.isoformat()
                if attempt.submitted_at
                else None,
            }
        )

    return {
        "summary": {
            "attempt_count": attempt_count,
            "average_score": _score_text(sum(scores) / attempt_count)
            if attempt_count
            else 0,
            "highest_score": _score_text(max(scores)) if scores else 0,
            "latest_score": _score_text(attempts[0].score) if attempts else 0,
            "weak_types": weak_types,
        },
        "type_stats": type_stats,
        "recent_attempts": recent_attempts,
    }


def _teacher_attempt_detail(attempt: models.MockAttempt):
    exam = attempt.mock_exam
    questions = sorted(exam.questions or [], key=lambda item: item.number) if exam else []
    answer_by_question_id = {
        answer.mock_question_id: answer for answer in attempt.answers or []
    }

    question_items = []
    for question in questions:
        answer = answer_by_question_id.get(question.id)
        selected_index = answer.selected_index if answer else None
        is_correct = bool(answer.is_correct) if answer else False
        question_items.append(
            {
                "question_id": question.id,
                "number": question.number,
                "question_type": question.question_type,
                "type_label": _type_label(question.question_type),
                "source": question.source,
                "passage": question.passage,
                "question_text": question.question_text,
                "options": question.options or [],
                "selected_index": selected_index,
                "answer_index": question.answer_index,
                "is_correct": is_correct,
                "explanation": question.explanation,
            }
        )

    weak_types = _weak_labels_for_attempt(attempt, questions)
    return {
        "student": {
            "user_id": attempt.user_id,
            "nickname": attempt.user.nickname
            if attempt.user
            else f"student{attempt.user_id}",
        },
        "attempt": {
            "id": attempt.id,
            "mock_exam_id": attempt.mock_exam_id,
            "title": exam.title if exam else "삭제된 모의고사",
            "grade": exam.grade if exam else "-",
            "year": exam.year if exam else None,
            "month": exam.month if exam else None,
            "score": _score_text(attempt.score),
            "correct_count": attempt.correct_count,
            "total_questions": attempt.total_questions,
            "submitted_at": attempt.submitted_at.isoformat()
            if attempt.submitted_at
            else None,
        },
        "summary": {
            "weak_types": weak_types,
            "correct_count": attempt.correct_count,
            "incorrect_count": max(
                (attempt.total_questions or 0) - (attempt.correct_count or 0),
                0,
            ),
            "score": _score_text(attempt.score),
        },
        "questions": question_items,
    }


def _mock_exam_report(exam: models.MockExam):
    questions = sorted(exam.questions or [], key=lambda item: item.number)
    latest_attempts = _latest_attempts_by_student(exam)
    attempt_count = len(latest_attempts)
    scores = [float(attempt.score or 0) for attempt in latest_attempts]

    question_type_counts: dict[str, int] = {}
    for question in questions:
        question_type_counts[question.question_type] = (
            question_type_counts.get(question.question_type, 0) + 1
        )

    correct_by_type: dict[str, int] = {type_code: 0 for type_code in TYPE_LABELS}
    total_by_type: dict[str, int] = {
        type_code: question_type_counts.get(type_code, 0) * attempt_count
        for type_code in TYPE_LABELS
    }
    for attempt in latest_attempts:
        for answer in attempt.answers or []:
            correct_by_type.setdefault(answer.question_type, 0)
            total_by_type.setdefault(answer.question_type, 0)
            if answer.is_correct:
                correct_by_type[answer.question_type] += 1

    ordered_types: list[str] = []
    for number in sorted(MOCK_TYPE_BY_NUMBER):
        type_code = MOCK_TYPE_BY_NUMBER[number]
        if type_code not in ordered_types:
            ordered_types.append(type_code)

    type_stats = []
    for type_code in ordered_types:
        total = total_by_type.get(type_code, 0)
        correct = correct_by_type.get(type_code, 0)
        rate = round(correct / total * 100, 1) if total else 0
        type_stats.append(
            {
                "type": type_code,
                "label": TYPE_LABELS.get(type_code, type_code),
                "correct": correct,
                "total": total,
                "rate": rate,
            }
        )

    students = []
    for attempt in sorted(
        latest_attempts,
        key=lambda item: item.submitted_at or item.started_at or datetime.min,
        reverse=True,
    ):
        students.append(
            {
                "user_id": attempt.user_id,
                "nickname": attempt.user.nickname if attempt.user else f"student{attempt.user_id}",
                "score": _score_text(attempt.score),
                "correct_count": attempt.correct_count,
                "total_questions": attempt.total_questions,
                "weak_types": _weak_labels_for_attempt(attempt, questions),
                "submitted_at": attempt.submitted_at.isoformat()
                if attempt.submitted_at
                else None,
            }
        )

    return {
        "mock_exam": {
            "id": exam.id,
            "title": exam.title,
            "grade": exam.grade,
            "year": exam.year,
            "month": exam.month,
            "total_questions": exam.total_questions,
            "question_count": len(questions),
            "is_complete": len(questions) == exam.total_questions,
        },
        "stats": {
            "attempt_count": attempt_count,
            "average_score": _score_text(sum(scores) / attempt_count)
            if attempt_count
            else 0,
            "highest_score": _score_text(max(scores)) if scores else 0,
            "lowest_score": _score_text(min(scores)) if scores else 0,
            "completion_rate": 100 if attempt_count else 0,
        },
        "type_stats": type_stats,
        "students": students,
    }


def _get_teacher_exam(db: Session, mock_exam_id: int, teacher_id: int):
    exam = (
        db.query(models.MockExam)
        .filter(
            models.MockExam.id == mock_exam_id,
            models.MockExam.created_by == teacher_id,
        )
        .first()
    )
    if not exam:
        raise HTTPException(status_code=404, detail="Mock exam not found")
    return exam


@router.post("")
def create_mock_exam(
    payload: MockExamCreateRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    _validate_exam_payload(payload)
    exam = models.MockExam(
        grade=payload.grade,
        year=payload.year,
        month=payload.month,
        title=payload.title,
        total_questions=20,
        total_score=100,
        has_listening=payload.has_listening,
        created_by=_user_id(current_user),
        created_at=datetime.utcnow(),
    )
    db.add(exam)
    db.commit()
    db.refresh(exam)
    return _serialize_exam(exam)


@router.get("")
def list_mock_exams(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exams = (
        db.query(models.MockExam)
        .filter(models.MockExam.created_by == _user_id(current_user))
        .order_by(
            models.MockExam.year.desc(),
            models.MockExam.month.desc(),
            models.MockExam.id.desc(),
        )
        .all()
    )
    return [_serialize_exam(exam) for exam in exams]


@router.get("/students/report")
def get_mock_exam_student_report_list(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    attempts = _teacher_mock_attempts(db, _user_id(current_user))
    by_student: dict[int, list[models.MockAttempt]] = {}
    for attempt in attempts:
        by_student.setdefault(attempt.user_id, []).append(attempt)

    students = []
    for user_id, student_attempts in by_student.items():
        report = _teacher_mock_student_report(student_attempts)
        summary = report["summary"]
        latest = student_attempts[0]
        students.append(
            {
                "user_id": user_id,
                "nickname": latest.user.nickname
                if latest.user
                else f"student{user_id}",
                "attempt_count": summary["attempt_count"],
                "average_score": summary["average_score"],
                "highest_score": summary["highest_score"],
                "latest_score": summary["latest_score"],
                "latest_submitted_at": latest.submitted_at.isoformat()
                if latest.submitted_at
                else None,
                "weak_types": summary["weak_types"],
            }
        )

    return {
        "students": sorted(
            students,
            key=lambda item: item["latest_submitted_at"] or "",
            reverse=True,
        )
    }


@router.get("/students/{student_id}/report")
def get_mock_exam_student_report_detail(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    attempts = _teacher_mock_attempts(db, _user_id(current_user), student_id)
    if not attempts:
        raise HTTPException(status_code=404, detail="Student mock report not found")

    report = _teacher_mock_student_report(attempts)
    user = attempts[0].user
    return {
        "student": {
            "user_id": student_id,
            "nickname": user.nickname if user else f"student{student_id}",
        },
        **report,
    }


@router.get("/students/{student_id}/attempts/{attempt_id}")
def get_mock_exam_student_attempt_detail(
    student_id: int,
    attempt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    attempt = (
        db.query(models.MockAttempt)
        .join(models.MockExam, models.MockAttempt.mock_exam_id == models.MockExam.id)
        .filter(
            models.MockAttempt.id == attempt_id,
            models.MockAttempt.user_id == student_id,
            models.MockExam.created_by == _user_id(current_user),
        )
        .first()
    )
    if not attempt:
        raise HTTPException(status_code=404, detail="Mock attempt not found")
    return _teacher_attempt_detail(attempt)


@router.get("/{mock_exam_id}")
def get_mock_exam(
    mock_exam_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam = _get_teacher_exam(db, mock_exam_id, _user_id(current_user))
    return _serialize_exam(exam, include_questions=True)


@router.get("/{mock_exam_id}/report")
def get_mock_exam_report(
    mock_exam_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam = _get_teacher_exam(db, mock_exam_id, _user_id(current_user))
    return _mock_exam_report(exam)


@router.delete("/{mock_exam_id}")
def delete_mock_exam(
    mock_exam_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam = _get_teacher_exam(db, mock_exam_id, _user_id(current_user))
    db.delete(exam)
    db.commit()
    return {"message": "Mock exam deleted", "mock_exam_id": mock_exam_id}


@router.post("/{mock_exam_id}/questions")
def upsert_mock_questions(
    mock_exam_id: int,
    payload: MockQuestionsUpsertRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam = _get_teacher_exam(db, mock_exam_id, _user_id(current_user))
    if not payload.questions:
        raise HTTPException(status_code=400, detail="questions are required")

    seen_numbers: set[int] = set()
    for question in payload.questions:
        _validate_question(question)
        if question.number in seen_numbers:
            raise HTTPException(
                status_code=400,
                detail=f"duplicate question number: {question.number}",
            )
        seen_numbers.add(question.number)

        passage_group_id = question.passage_group_id

        existing = (
            db.query(models.MockQuestion)
            .filter(
                models.MockQuestion.mock_exam_id == exam.id,
                models.MockQuestion.number == question.number,
            )
            .first()
        )
        if existing:
            existing.question_type = question.question_type
            existing.source = question.source
            existing.passage = question.passage
            existing.question_text = question.question_text
            existing.options = question.options
            existing.answer_index = question.answer_index
            existing.explanation = question.explanation
            existing.passage_group_id = passage_group_id
        else:
            db.add(
                models.MockQuestion(
                    mock_exam_id=exam.id,
                    number=question.number,
                    question_type=question.question_type,
                    source=question.source,
                    passage=question.passage,
                    question_text=question.question_text,
                    options=question.options,
                    answer_index=question.answer_index,
                    explanation=question.explanation,
                    passage_group_id=passage_group_id,
                    created_at=datetime.utcnow(),
                )
            )

    db.commit()
    db.refresh(exam)
    return _serialize_exam(exam, include_questions=True)


@router.post("/{mock_exam_id}/questions/upload")
async def upload_mock_questions(
    mock_exam_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam = _get_teacher_exam(db, mock_exam_id, _user_id(current_user))
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    rows = _read_upload_rows(file.filename or "", raw)
    questions = _questions_from_upload_rows(rows)

    for question in questions:
        _validate_question(question)
        passage_group_id = question.passage_group_id

        existing = (
            db.query(models.MockQuestion)
            .filter(
                models.MockQuestion.mock_exam_id == exam.id,
                models.MockQuestion.number == question.number,
            )
            .first()
        )
        if existing:
            existing.question_type = question.question_type
            existing.source = question.source
            existing.passage = question.passage
            existing.question_text = question.question_text
            existing.options = question.options
            existing.answer_index = question.answer_index
            existing.explanation = question.explanation
            existing.passage_group_id = passage_group_id
        else:
            db.add(
                models.MockQuestion(
                    mock_exam_id=exam.id,
                    number=question.number,
                    question_type=question.question_type,
                    source=question.source,
                    passage=question.passage,
                    question_text=question.question_text,
                    options=question.options,
                    answer_index=question.answer_index,
                    explanation=question.explanation,
                    passage_group_id=passage_group_id,
                    created_at=datetime.utcnow(),
                )
            )

    db.commit()
    db.refresh(exam)
    return {
        "message": "Mock exam questions uploaded",
        "mock_exam": _serialize_exam(exam, include_questions=True),
    }


@router.patch("/{mock_exam_id}/questions/{question_id}")
def update_mock_question(
    mock_exam_id: int,
    question_id: int,
    payload: MockQuestionPatchRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    _, question = _get_teacher_question(
        db,
        mock_exam_id,
        question_id,
        _user_id(current_user),
    )

    if payload.options is not None:
        cleaned_options = [_clean_cell(option) for option in payload.options]
        if len(cleaned_options) != 5 or any(not option for option in cleaned_options):
            raise HTTPException(
                status_code=400,
                detail="options must contain exactly 5 non-empty items",
            )
        question.options = cleaned_options

    selected_answer = payload.answer_index
    if payload.answer is not None:
        if payload.answer < 1 or payload.answer > 5:
            raise HTTPException(status_code=400, detail="answer must be 1-5")
        selected_answer = payload.answer - 1
    if selected_answer is not None:
        if selected_answer < 0 or selected_answer > 4:
            raise HTTPException(status_code=400, detail="answer_index must be 0-4")
        question.answer_index = selected_answer

    if payload.source is not None:
        question.source = _clean_cell(payload.source) or None
    if payload.passage is not None:
        question.passage = payload.passage
    if payload.question_text is not None:
        question_text = _clean_cell(payload.question_text)
        if not question_text:
            raise HTTPException(status_code=400, detail="question_text is required")
        question.question_text = question_text
    if payload.explanation is not None:
        question.explanation = payload.explanation
    if payload.passage_group_id is not None:
        question.passage_group_id = _clean_cell(payload.passage_group_id) or None

    db.commit()
    db.refresh(question)
    return _serialize_question(question, include_answer=True)


@router.delete("/{mock_exam_id}/questions/{question_id}")
def delete_mock_question(
    mock_exam_id: int,
    question_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("teacher")),
):
    exam, question = _get_teacher_question(
        db,
        mock_exam_id,
        question_id,
        _user_id(current_user),
    )
    db.delete(question)
    db.commit()
    db.refresh(exam)
    return {"message": "Mock question deleted", "mock_exam": _serialize_exam(exam)}
